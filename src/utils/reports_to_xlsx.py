from openpyxl.styles import PatternFill, Font, Alignment
import pandas as pd

def reports_to_xlsx(reports):
    """
    Convert a list of reports to an Excel file.

    Args:
        reports (list): List of dictionaries representing the reports.

    Returns:
        str: Path to the generated Excel file.
    """
    if not reports:
        return None

    df = pd.DataFrame([dict(r) for r in reports])

    def define_type(row):
        if row.get("is_plan"):
            return "План"
        elif row.get("is_report"):
            return "Звіт"
        else:
            return "—"

    df["Тип"] = df.apply(define_type, axis=1)

    columns_order = [
        "name",
        "ads",
        "responses",
        "records",
        "Тип",
        "created_at"
    ]

    df = df[columns_order]

    df.columns = [
        "Користувач",
        "Оголошення",
        "Відповіді",
        "Записи",
        "Тип",
        "Дата створення"
    ]

    df["Дата створення"] = pd.to_datetime(df["Дата створення"]).dt.strftime('%d.%m.%Y %H:%M')

    file_path = 'reports.xlsx'
    with pd.ExcelWriter(file_path, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name='Reports')

        workbook = writer.book
        worksheet = workbook['Reports']

        widths = {
            "A": 15,
            "B": 15,
            "C": 10,
            "D": 10,
            "E": 10,
            "F": 15,
        }

        for col, width in widths.items():
            worksheet.column_dimensions[col].width = width

        header_font = Font(bold=True, color='FFFFFF')
        header_fill = PatternFill(start_color='4F81BD', end_color='4F81BD', fill_type='solid')
        for cell in worksheet[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center', vertical='center')

        for row in worksheet.iter_rows(min_row=2, max_row=worksheet.max_row):
            for cell in row:
                cell.alignment = Alignment(horizontal='left', vertical='center')

    return file_path
