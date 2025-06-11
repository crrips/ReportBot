from openpyxl.styles import PatternFill, Font, Alignment
import pandas as pd

def records_to_xlsx(records):
    """
    Convert a list of records to an Excel file.

    Args:
        records (list): List of dictionaries representing the records.

    Returns:
        str: Path to the generated Excel file.
    """
    if not records:
        return None

    df = pd.DataFrame([dict(r) for r in records])

    file_path = 'records.xlsx'

    columns_order = [
        "name",
        "student",
        "subject",
        "link",
        "phone_number",
        "lesson_date",
        "lesson_time",
        "created_at"
    ]
    
    df = df[columns_order]
    
    df.columns = [
        "Користувач",
        "Студент",
        "Предмет",
        "Посилання",
        "Номер телефону",
        "Дата заняття",
        "Час заняття",
        "Дата створення"
    ]
    
    df["Дата заняття"] = pd.to_datetime(df["Дата заняття"]).dt.strftime('%d.%m.%Y')
    df["Час заняття"] = df["Час заняття"].apply(lambda t: t.strftime('%H:%M'))
    df["Дата створення"] = pd.to_datetime(df["Дата створення"]).dt.strftime('%d.%m.%Y %H:%M')
    
    with pd.ExcelWriter(file_path, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name='Records')

        workbook = writer.book
        worksheet = workbook['Records']
        
        widths = {
            "A": 15,
            "B": 20,
            "C": 15,
            "D": 20,
            "E": 20,
            "F": 15,
            "G": 15,
            "H": 15
        }
        for col, width in widths.items():
            worksheet.column_dimensions[col].width = width

        header_font = Font(bold=True, color='FFFFFF')
        header_fill = PatternFill(start_color='4F81BD', end_color='4F81BD', fill_type='solid')
        for cell in worksheet[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center', vertical='center')
            
        for row in worksheet.iter_rows(min_row=2, max_row=worksheet.max_row):
            for cell in row:
                cell.alignment = Alignment(horizontal='left', vertical='center')
    
    return file_path